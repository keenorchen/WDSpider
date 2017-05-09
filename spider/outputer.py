from openpyxl import Workbook, utils
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import Dimension, ColumnDimension


class Outputer(object):
    def __init__(self):
        self.data_sheet1 = []
        self.data_sheet2 = []

    wdzj_title = [
        ("platName", "平台"),
        ("amount", "成交量(万元)"),
        ("bidderNum", "投资人数"),
        ("borrowerNum", "借款人数"),
        ("incomeRate", "平均收益率"),
        ("loanPeriod", "平均借款期限(月)"),
        ("stayStillOfTotal", "待还余额"),
        ("netInflowOfThirty", "资金净流入"),
        ("timeOperation", "运营时间(月)")
    ]

    def collect_data(self, data, data2=None):
        if data is None:
            return
        self.data_sheet1 = data
        self.data_sheet2 = data2

    def output_html(self):
        pass

    def output_json(self, file_name):
        wb = Workbook()
        sheet = wb.get_active_sheet()
        if not sheet:
            sheet = wb.create_sheet()

        sheet.title = '网贷之家'
        sheet2 = wb.create_sheet(title='网贷天眼')


        # 写入列标题
        for i, value in enumerate(self.wdzj_title):
            # print(str(i) +'  ' + value[1])
            sheet.cell(row=1, column=i + 1).value = value[1]

        # 写入数据
        for i, value in enumerate(self.data_sheet1):
            for j in range(0, len(self.wdzj_title)):
                sheet.cell(row=i + 2, column=j + 1).value = value[self.wdzj_title[j][0]]

        for row in self.data_sheet2:
            sheet2.append(row)

        # 设置Excel样式
        self.set_style(sheet)
        self.set_style(sheet2)

        # 冻结第一行
        sheet.freeze_panes = 'A2'
        sheet2.freeze_panes = 'A2'

        # 自适应列宽调整
        self.fit_colnum_width(sheet)
        self.fit_colnum_width(sheet2)

        wb.save("../output_file/spider_data_" + file_name + ".xlsx")

    def fit_colnum_width(self, sheet):
        column_widths = []
        for row in sheet.rows:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell)) > column_widths[i]:
                        column_widths[i] = len(str(cell))
                else:
                    column_widths += [len(str(cell))]
        for i, column_width in enumerate(column_widths):
            sheet.column_dimensions[utils.get_column_letter(i + 1)].width = column_width

    def set_style(self, sheet):
        font = Font(size=14)
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                sheet.cell(row=i, column=j).font = font
